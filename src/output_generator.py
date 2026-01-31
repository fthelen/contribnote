"""
Output Generator Module

Creates the output Excel workbook with commentary results.
"""

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from .selection_engine import SelectionResult, RankedSecurity, SelectionMode
from .openai_client import CommentaryResult, Citation


@dataclass
class OutputRow:
    """A single row in the output workbook."""
    ticker: str
    security_name: str
    rank: Optional[int]
    contributor_detractor: str
    contribution_to_return: float
    port_ending_weight: float
    commentary: str
    sources: str
    is_error: bool = False


def format_citations(citations: list[Citation]) -> str:
    """
    Format citations as numbered list.
    
    Example output:
    [1] https://example.com/article1
    [2] https://example.com/article2
    """
    if not citations:
        return ""
    
    lines = []
    for i, citation in enumerate(citations, 1):
        lines.append(f"[{i}] {citation.url}")
    
    return "\n".join(lines)


def merge_results(
    selection: SelectionResult,
    commentary_results: dict[str, CommentaryResult]
) -> list[OutputRow]:
    """
    Merge selection results with commentary results.
    
    Args:
        selection: Selection result for a portfolio
        commentary_results: Dict mapping ticker to CommentaryResult
        
    Returns:
        List of OutputRow objects ready for export
    """
    rows = []
    
    for ranked_sec in selection.ranked_securities:
        ticker = ranked_sec.ticker
        result = commentary_results.get(ticker)
        
        if result and result.success:
            commentary = result.commentary
            sources = format_citations(result.citations)
            is_error = False
        elif result:
            commentary = f"ERROR: {result.error_message}"
            sources = ""
            is_error = True
        else:
            commentary = "ERROR: No commentary generated"
            sources = ""
            is_error = True
        
        rows.append(OutputRow(
            ticker=ticker,
            security_name=ranked_sec.security_name,
            rank=ranked_sec.rank,
            contributor_detractor=ranked_sec.security_type.value,
            contribution_to_return=ranked_sec.contribution_to_return,
            port_ending_weight=ranked_sec.port_ending_weight,
            commentary=commentary,
            sources=sources,
            is_error=is_error
        ))
    
    return rows


def create_output_workbook(
    selections: list[SelectionResult],
    commentary_results: dict[str, dict[str, CommentaryResult]],
    output_folder: Path
) -> Path:
    """
    Create the output Excel workbook.
    
    Args:
        selections: List of SelectionResult objects (one per portfolio)
        commentary_results: Nested dict: portcode -> ticker -> CommentaryResult
        output_folder: Folder to save the output file
        
    Returns:
        Path to the created workbook
    """
    # Generate filename
    timestamp = datetime.now().strftime("%Y-%m-%d_%H%M")
    filename = f"ContributorDetractorCommentary_{timestamp}.xlsx"
    output_path = output_folder / filename
    
    # Ensure output folder exists
    output_folder.mkdir(parents=True, exist_ok=True)
    
    # Create workbook
    wb = openpyxl.Workbook()
    if wb.active is not None:
        wb.remove(wb.active)  # Remove default sheet
    
    # Define styles
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    wrap_alignment = Alignment(wrap_text=True, vertical='top')
    
    # Column headers
    headers = [
        "Ticker",
        "Security Name",
        "Rank",
        "Contributor/Detractor",
        "Contribution To Return",
        "Port. Ending Weight",
        "Commentary",
        "Sources"
    ]
    
    # Column widths
    column_widths = [12, 30, 8, 18, 20, 18, 60, 40]
    
    for selection in selections:
        # Create sheet (trim name to 31 chars max for Excel)
        sheet_name = selection.portcode[:31]
        ws = wb.create_sheet(title=sheet_name)
        
        # Get commentary results for this portfolio
        port_commentary = commentary_results.get(selection.portcode, {})
        
        # Merge results
        rows = merge_results(selection, port_commentary)
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Set column widths
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # Write data rows
        for row_num, output_row in enumerate(rows, 2):
            # Ticker
            ws.cell(row=row_num, column=1, value=output_row.ticker).border = thin_border
            
            # Security Name
            ws.cell(row=row_num, column=2, value=output_row.security_name).border = thin_border
            
            # Rank
            rank_cell = ws.cell(row=row_num, column=3, value=output_row.rank if output_row.rank else "")
            rank_cell.border = thin_border
            rank_cell.alignment = Alignment(horizontal='center')
            
            # Contributor/Detractor
            ws.cell(row=row_num, column=4, value=output_row.contributor_detractor).border = thin_border
            
            # Contribution To Return (formatted to 4 decimal places as percentage)
            contrib_cell = ws.cell(row=row_num, column=5, value=output_row.contribution_to_return)
            contrib_cell.number_format = '0.0000'
            contrib_cell.border = thin_border
            
            # Port. Ending Weight (formatted to 2 decimal places)
            weight_cell = ws.cell(row=row_num, column=6, value=output_row.port_ending_weight)
            weight_cell.number_format = '0.00'
            weight_cell.border = thin_border
            
            # Commentary
            commentary_cell = ws.cell(row=row_num, column=7, value=output_row.commentary)
            commentary_cell.border = thin_border
            commentary_cell.alignment = wrap_alignment
            if output_row.is_error:
                commentary_cell.font = Font(color="FF0000")
            
            # Sources
            sources_cell = ws.cell(row=row_num, column=8, value=output_row.sources)
            sources_cell.border = thin_border
            sources_cell.alignment = wrap_alignment
        
        # Freeze header row
        ws.freeze_panes = 'A2'
    
    # Save workbook
    wb.save(output_path)
    
    return output_path


def create_log_file(
    output_folder: Path,
    input_files: list[Path],
    output_file: Path,
    errors: dict[str, list[str]],
    start_time: datetime,
    end_time: datetime
) -> Path:
    """
    Create a log file for the run.
    
    Args:
        output_folder: Output folder
        input_files: List of input file paths
        output_file: Path to the output workbook
        errors: Dict mapping "PORTCODE|TICKER" to error messages
        start_time: Run start time
        end_time: Run end time
        
    Returns:
        Path to the log file
    """
    # Create log folder
    log_folder = output_folder / "log"
    log_folder.mkdir(parents=True, exist_ok=True)
    
    # Generate log filename
    timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    log_path = log_folder / f"run_log_{timestamp}.txt"
    
    with open(log_path, 'w') as f:
        f.write("=" * 60 + "\n")
        f.write("Commentary Generator - Run Log\n")
        f.write("=" * 60 + "\n\n")
        
        f.write(f"Run Timestamp: {start_time.strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write(f"Duration: {(end_time - start_time).total_seconds():.1f} seconds\n\n")
        
        f.write("Input Files Processed:\n")
        for file_path in input_files:
            f.write(f"  - {file_path.name}\n")
        f.write("\n")
        
        f.write(f"Output Workbook: {output_file.name}\n\n")
        
        if errors:
            f.write("Errors:\n")
            for key, error_list in errors.items():
                for error in error_list:
                    f.write(f"  [{key}] {error}\n")
        else:
            f.write("No errors encountered.\n")
        
        f.write("\n" + "=" * 60 + "\n")
    
    return log_path
