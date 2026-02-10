"""
Excel Parser Module

Reads FactSet Excel files and extracts portfolio data.
- Tab: ContributionMasterRisk
- Header row: 7
- Data starts: row 10
- Period string: row 6
- End marker: first blank Ticker cell
"""

from dataclasses import dataclass
from pathlib import Path
from typing import Optional

import openpyxl


@dataclass
class SecurityRow:
    """Represents a single security row from the Excel file."""
    ticker: str
    security_name: str
    port_ending_weight: float
    contribution_to_return: float
    gics: Optional[str]
    
    def is_cash_or_fee(self) -> bool:
        """Check if this row is cash or fees (GICS == 'NA' or dash markers)."""
        if self.gics is None:
            return True
        return self.gics in {"NA", "—", "--"}


@dataclass
class PortfolioData:
    """Parsed data from a single portfolio Excel file."""
    portcode: str
    period: str
    securities: list[SecurityRow]
    source_file: Path
    sector_attribution: Optional["AttributionTable"] = None
    country_attribution: Optional["AttributionTable"] = None
    attribution_warnings: Optional[list[str]] = None
    
    def get_filtered_securities(self) -> list[SecurityRow]:
        """Return securities excluding cash/fees rows."""
        return [s for s in self.securities if not s.is_cash_or_fee()]

    def __post_init__(self) -> None:
        """Initialize mutable defaults safely."""
        if self.attribution_warnings is None:
            self.attribution_warnings = []


@dataclass
class AttributionRow:
    """A single top-level attribution row."""
    category: str
    metrics: dict[str, float | str]


@dataclass
class AttributionTable:
    """Parsed attribution table from a workbook sheet."""
    sheet_name: str
    category_header: str
    metric_headers: list[str]
    top_level_rows: list[AttributionRow]
    total_row: Optional[AttributionRow]

    def has_data(self) -> bool:
        """Return True when the table has usable attribution content."""
        return bool(self.top_level_rows or self.total_row)


ATTRIBUTION_CATEGORY_HEADERS = {
    "AttributionbyCountryMasterRisk": "Country",
    "AttributionbySector": "Sector",
}


def extract_portcode_from_filename(filename: str) -> str:
    """
    Extract PORTCODE from filename.
    Pattern: PORTCODE_*_MMDDYYYY.xlsx
    PORTCODE = everything before the first underscore
    """
    base_name = Path(filename).stem  # Remove .xlsx
    parts = base_name.split('_')
    if parts:
        return parts[0]
    return base_name


def _parse_numeric_or_text(value: object) -> float | str:
    """Coerce numeric-looking values to float, otherwise return cleaned text."""
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return ""
    try:
        return float(text.replace(",", ""))
    except ValueError:
        return text


def _parse_metric_headers(ws: openpyxl.worksheet.worksheet.Worksheet) -> list[str]:
    """Parse attribution metric headers from row 7, columns B+."""
    headers: list[str] = []
    for col in range(2, ws.max_column + 1):
        value = ws.cell(row=7, column=col).value
        if value is None:
            continue
        header = str(value).strip()
        if header:
            headers.append(header)
    return headers


def _build_attribution_row(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    row_num: int,
    category: str,
    metric_headers: list[str]
) -> AttributionRow:
    """Build an AttributionRow from a worksheet row."""
    metrics: dict[str, float | str] = {}
    for offset, header in enumerate(metric_headers, start=2):
        value = ws.cell(row=row_num, column=offset).value
        metrics[header] = _parse_numeric_or_text(value)
    return AttributionRow(category=category, metrics=metrics)


def _parse_attribution_sheet(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    sheet_name: str,
    file_path: Path,
    warnings: list[str]
) -> Optional[AttributionTable]:
    """Parse an attribution sheet, keeping only highest-level grouped rows."""
    metric_headers = _parse_metric_headers(ws)
    if not metric_headers:
        warnings.append(
            f"{file_path.name} [{sheet_name}]: Missing metric headers in row 7; "
            "attribution data skipped."
        )
        return None

    category_header = ATTRIBUTION_CATEGORY_HEADERS.get(sheet_name, "Category")
    candidate_rows: list[tuple[int, AttributionRow]] = []
    total_row: Optional[AttributionRow] = None

    seen_data = False
    consecutive_blank_categories = 0
    for row_num in range(8, ws.max_row + 1):
        category_raw = ws.cell(row=row_num, column=1).value
        if category_raw is None:
            if seen_data:
                consecutive_blank_categories += 1
                # Allow occasional blank category rows after data starts (e.g., before "Total")
                if consecutive_blank_categories >= 3:
                    break
            continue

        category = str(category_raw).strip()
        if not category:
            if seen_data:
                consecutive_blank_categories += 1
                # Allow occasional blank category rows after data starts (e.g., before "Total")
                if consecutive_blank_categories >= 3:
                    break
            continue

        seen_data = True
        consecutive_blank_categories = 0

        row_dim = ws.row_dimensions.get(row_num)
        outline_level = row_dim.outlineLevel if row_dim is not None else 0
        if outline_level is None:
            outline_level = 0
        row = _build_attribution_row(ws, row_num, category, metric_headers)

        if category.lower() == "total":
            total_row = row
            continue

        candidate_rows.append((outline_level, row))

    if candidate_rows:
        top_level_outline = min(outline for outline, _ in candidate_rows)
        top_level_rows = [
            row for outline, row in candidate_rows if outline == top_level_outline
        ]
    else:
        top_level_rows = []

    if total_row is None:
        warnings.append(
            f"{file_path.name} [{sheet_name}]: Total row not found; parsed top-level "
            "rows without total."
        )

    if not top_level_rows:
        warnings.append(
            f"{file_path.name} [{sheet_name}]: No top-level attribution rows were found."
        )

    table = AttributionTable(
        sheet_name=sheet_name,
        category_header=category_header,
        metric_headers=metric_headers,
        top_level_rows=top_level_rows,
        total_row=total_row,
    )

    if not table.has_data():
        return None

    return table


def _format_markdown_metric(value: float | str) -> str:
    """Format an attribution metric value for markdown output."""
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return f"{value:.6g}"
    return str(value)


def format_attribution_table_markdown(
    table: Optional[AttributionTable],
    empty_message: str
) -> str:
    """
    Format an attribution table for prompt injection as markdown.

    Includes top-level rows with an optional Total row at the bottom.
    """
    if table is None or not table.has_data():
        return empty_message

    headers = [table.category_header] + table.metric_headers
    lines: list[str] = []
    lines.append(f"### {table.sheet_name}")
    lines.append("")
    lines.append("| " + " | ".join(headers) + " |")
    lines.append("| " + " | ".join(["---"] * len(headers)) + " |")

    for row in table.top_level_rows:
        values = [row.category] + [
            _format_markdown_metric(row.metrics.get(header, ""))
            for header in table.metric_headers
        ]
        lines.append("| " + " | ".join(values) + " |")

    if table.total_row is not None:
        total_values = [table.total_row.category] + [
            _format_markdown_metric(table.total_row.metrics.get(header, ""))
            for header in table.metric_headers
        ]
        lines.append("| " + " | ".join(total_values) + " |")

    return "\n".join(lines)


def parse_excel_file(file_path: Path) -> PortfolioData:
    """
    Parse a FactSet Excel file and extract portfolio data.
    
    Args:
        file_path: Path to the Excel file
        
    Returns:
        PortfolioData object with extracted information
        
    Raises:
        ValueError: If the file format is invalid or required data is missing
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    attribution_warnings: list[str] = []
    
    # Check for required sheet
    sheet_name = "ContributionMasterRisk"
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Required sheet '{sheet_name}' not found in {file_path}")
    
    ws = wb[sheet_name]
    
    # Extract period from row 6
    period_value = ws.cell(row=6, column=1).value
    if not period_value:
        raise ValueError(f"Period not found in row 6 of {file_path}")
    period = str(period_value).strip()
    
    # Find column indices by scanning row 7 for headers
    # Security Name header may be blank, but it's always left of Ticker
    col_map = {}
    ticker_col = None
    
    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=7, column=col).value
        if header is None:
            continue
        header_str = str(header).strip()
        if not header_str:
            continue
        header_key = header_str.lower()
        if header_key == "ticker":
            col_map["Ticker"] = col
            ticker_col = col
        elif header_key == "security name":
            col_map["Security Name"] = col
        elif header_key == "port. ending weight":
            col_map["Port. Ending Weight"] = col
        elif header_key == "contribution to return":
            col_map["Contribution To Return"] = col
        elif header_key == "gics":
            col_map["GICS"] = col
    
    # Validate required columns found
    required_cols = ["Ticker", "Port. Ending Weight", "Contribution To Return", "GICS"]
    missing = [c for c in required_cols if c not in col_map]
    if missing:
        raise ValueError(f"Missing required columns in {file_path}: {missing}")
    
    # If Security Name header not found, use column left of Ticker
    if "Security Name" not in col_map and ticker_col and ticker_col > 1:
        col_map["Security Name"] = ticker_col - 1
    elif "Security Name" not in col_map:
        col_map["Security Name"] = None  # Will default to empty string
    
    # Parse data rows starting at row 10
    securities = []
    row_num = 10
    
    while True:
        ticker = ws.cell(row=row_num, column=col_map["Ticker"]).value
        
        # Stop at first blank ticker (end of table)
        if ticker is None or str(ticker).strip() == "":
            break
        
        security_name = ws.cell(row=row_num, column=col_map["Security Name"]).value if col_map["Security Name"] else ""
        port_ending_weight = ws.cell(row=row_num, column=col_map["Port. Ending Weight"]).value
        contribution_to_return = ws.cell(row=row_num, column=col_map["Contribution To Return"]).value
        gics = ws.cell(row=row_num, column=col_map["GICS"]).value
        
        # Parse numeric values (handle potential None or string values)
        try:
            weight = float(port_ending_weight) if port_ending_weight is not None else 0.0
        except (ValueError, TypeError):
            weight = 0.0
            
        try:
            contribution = float(contribution_to_return) if contribution_to_return is not None else 0.0
        except (ValueError, TypeError):
            contribution = 0.0
        
        securities.append(SecurityRow(
            ticker=str(ticker).strip(),
            security_name=str(security_name).strip() if security_name else "",
            port_ending_weight=weight,
            contribution_to_return=contribution,
            gics=str(gics).strip() if gics else "NA"
        ))
        
        row_num += 1
    
    # Parse optional attribution tabs (exact names only)
    sector_sheet_name = "AttributionbySector"
    country_sheet_name = "AttributionbyCountryMasterRisk"

    if sector_sheet_name in wb.sheetnames:
        sector_attribution = _parse_attribution_sheet(
            wb[sector_sheet_name], sector_sheet_name, file_path, attribution_warnings
        )
    else:
        sector_attribution = None
        attribution_warnings.append(
            f"{file_path.name}: Missing expected attribution tab '{sector_sheet_name}'."
        )

    if country_sheet_name in wb.sheetnames:
        country_attribution = _parse_attribution_sheet(
            wb[country_sheet_name], country_sheet_name, file_path, attribution_warnings
        )
    else:
        country_attribution = None
        attribution_warnings.append(
            f"{file_path.name}: Missing optional attribution tab '{country_sheet_name}'."
        )

    wb.close()
    
    # Extract portcode from filename
    portcode = extract_portcode_from_filename(file_path.name)
    
    return PortfolioData(
        portcode=portcode,
        period=period,
        securities=securities,
        source_file=file_path,
        sector_attribution=sector_attribution,
        country_attribution=country_attribution,
        attribution_warnings=attribution_warnings,
    )


def parse_multiple_files(file_paths: list[Path]) -> list[PortfolioData]:
    """
    Parse multiple Excel files.
    
    Args:
        file_paths: List of paths to Excel files
        
    Returns:
        List of PortfolioData objects
    """
    results = []
    for file_path in file_paths:
        try:
            data = parse_excel_file(file_path)
            results.append(data)
        except Exception as e:
            print(f"Error parsing {file_path}: {e}")
            raise
    return results
