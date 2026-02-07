"""
Tests for the Output Generator Module.
"""
import pytest
from pathlib import Path
from datetime import datetime
from unittest.mock import MagicMock, patch
import tempfile
import os

from src.selection_engine import (
    SelectionResult,
    SelectionMode,
    RankedSecurity,
    SecurityType,
)
from src.excel_parser import SecurityRow
from src.openai_client import CommentaryResult, Citation, AttributionOverviewResult
from src.output_generator import (
    OutputRow,
    format_citations,
    merge_results,
    create_output_workbook,
    create_log_file,
)


# --- Helper Functions ---

def make_security(ticker: str, contribution: float, weight: float = 1.0) -> SecurityRow:
    """Helper to create a SecurityRow for testing."""
    return SecurityRow(
        ticker=ticker,
        security_name=f"{ticker} Inc.",
        port_ending_weight=weight,
        contribution_to_return=contribution,
        gics="Technology"
    )


def make_ranked_security(
    ticker: str,
    contribution: float,
    rank: int = 1,
    security_type: SecurityType = SecurityType.CONTRIBUTOR
) -> RankedSecurity:
    """Helper to create a RankedSecurity for testing."""
    return RankedSecurity(
        security=make_security(ticker, contribution),
        rank=rank,
        security_type=security_type
    )


def make_selection_result(
    ranked_securities: list[RankedSecurity],
    portcode: str = "TEST"
) -> SelectionResult:
    """Helper to create a SelectionResult for testing."""
    return SelectionResult(
        portcode=portcode,
        period="12/31/2025 to 1/28/2026",
        ranked_securities=ranked_securities,
        mode=SelectionMode.TOP_BOTTOM,
        source_file=f"{portcode}_12312025_01282026.xlsx"
    )


def make_commentary_result(
    ticker: str,
    commentary: str = "Test commentary.",
    citations: list[Citation] = None,
    success: bool = True,
    error_message: str = ""
) -> CommentaryResult:
    """Helper to create a CommentaryResult for testing."""
    return CommentaryResult(
        ticker=ticker,
        security_name=f"{ticker} Inc.",
        commentary=commentary,
        citations=citations or [],
        success=success,
        error_message=error_message
    )


def make_attribution_overview_result(
    portcode: str,
    output: str = "Attribution overview.",
    citations: list[Citation] = None,
    success: bool = True,
    error_message: str = "",
) -> AttributionOverviewResult:
    """Helper to create an AttributionOverviewResult for testing."""
    return AttributionOverviewResult(
        portcode=portcode,
        output=output,
        citations=citations or [],
        success=success,
        error_message=error_message,
    )


# --- format_citations Tests ---

class TestFormatCitations:
    """Tests for the format_citations function."""

    def test_format_single_citation(self):
        """Should format a single citation."""
        citations = [Citation(url="https://example.com/article")]
        
        result = format_citations(citations)
        
        assert result == "[1] https://example.com/article"

    def test_format_multiple_citations(self):
        """Should format multiple citations with numbering."""
        citations = [
            Citation(url="https://example.com/article1"),
            Citation(url="https://example.com/article2"),
            Citation(url="https://example.com/article3"),
        ]
        
        result = format_citations(citations)
        
        lines = result.split("\n")
        assert len(lines) == 3
        assert lines[0] == "[1] https://example.com/article1"
        assert lines[1] == "[2] https://example.com/article2"
        assert lines[2] == "[3] https://example.com/article3"

    def test_format_empty_citations(self):
        """Should return empty string for no citations."""
        result = format_citations([])
        
        assert result == ""

    def test_format_citation_with_title(self):
        """Should handle citations with titles (but not include title in output)."""
        citations = [Citation(url="https://example.com", title="Example Article")]
        
        result = format_citations(citations)
        
        # Current implementation only includes URL
        assert "[1] https://example.com" in result


# --- merge_results Tests ---

class TestMergeResults:
    """Tests for the merge_results function."""

    def test_merge_successful_results(self):
        """Should merge successful commentary results."""
        ranked = [
            make_ranked_security("AAPL", 0.15, rank=1),
            make_ranked_security("MSFT", 0.10, rank=2),
        ]
        selection = make_selection_result(ranked)
        
        commentary = {
            "AAPL": make_commentary_result("AAPL", "Apple did great."),
            "MSFT": make_commentary_result("MSFT", "Microsoft performed well."),
        }
        
        rows = merge_results(selection, commentary)
        
        assert len(rows) == 2
        assert rows[0].ticker == "AAPL"
        assert rows[0].commentary == "Apple did great."
        assert rows[0].is_error is False
        assert rows[1].ticker == "MSFT"
        assert rows[1].commentary == "Microsoft performed well."

    def test_merge_with_citations(self):
        """Should include formatted citations in sources column."""
        ranked = [make_ranked_security("AAPL", 0.15)]
        selection = make_selection_result(ranked)
        
        citations = [
            Citation(url="https://reuters.com/article1"),
            Citation(url="https://bloomberg.com/article2"),
        ]
        commentary = {"AAPL": make_commentary_result("AAPL", "Commentary.", citations)}
        
        rows = merge_results(selection, commentary)
        
        assert "[1] https://reuters.com/article1" in rows[0].sources
        assert "[2] https://bloomberg.com/article2" in rows[0].sources

    def test_merge_with_error_result(self):
        """Should handle error results."""
        ranked = [make_ranked_security("AAPL", 0.15)]
        selection = make_selection_result(ranked)
        
        commentary = {
            "AAPL": make_commentary_result(
                "AAPL",
                commentary="",
                success=False,
                error_message="API timeout"
            )
        }
        
        rows = merge_results(selection, commentary)
        
        assert rows[0].is_error is True
        assert "ERROR: API timeout" in rows[0].commentary
        assert rows[0].sources == ""

    def test_merge_with_missing_result(self):
        """Should handle missing commentary results."""
        ranked = [make_ranked_security("AAPL", 0.15)]
        selection = make_selection_result(ranked)
        
        commentary = {}  # No results
        
        rows = merge_results(selection, commentary)
        
        assert rows[0].is_error is True
        assert "ERROR: No commentary generated" in rows[0].commentary

    def test_merge_preserves_security_data(self):
        """Should preserve all security metadata."""
        security = make_security("AAPL", 0.15, weight=5.25)
        ranked = [RankedSecurity(security, rank=1, security_type=SecurityType.CONTRIBUTOR)]
        selection = make_selection_result(ranked)
        
        commentary = {"AAPL": make_commentary_result("AAPL", "Commentary.")}
        
        rows = merge_results(selection, commentary)
        
        assert rows[0].ticker == "AAPL"
        assert rows[0].security_name == "AAPL Inc."
        assert rows[0].rank == 1
        assert rows[0].contributor_detractor == "Contributor"
        assert rows[0].contribution_to_return == 0.15
        assert rows[0].port_ending_weight == 5.25


# --- create_output_workbook Tests ---

class TestCreateOutputWorkbook:
    """Tests for the create_output_workbook function."""

    def test_creates_workbook_file(self):
        """Should create an Excel workbook file."""
        with tempfile.TemporaryDirectory() as tmpdir:
            output_folder = Path(tmpdir)
            
            ranked = [make_ranked_security("AAPL", 0.15)]
            selection = make_selection_result(ranked, portcode="TESTPORT")
            commentary = {"TESTPORT": {"AAPL": make_commentary_result("AAPL", "Commentary.")}}
            
            result_path = create_output_workbook([selection], commentary, output_folder)
            
            assert result_path.exists()
            assert result_path.suffix == ".xlsx"
            assert "ContributorDetractorCommentary_" in result_path.name

    def test_creates_sheet_per_portfolio(self):
        """Should create one sheet per portfolio."""
        import openpyxl
        
        with tempfile.TemporaryDirectory() as tmpdir:
            output_folder = Path(tmpdir)
            
            selections = [
                make_selection_result([make_ranked_security("A", 0.1)], portcode="PORT1"),
                make_selection_result([make_ranked_security("B", 0.2)], portcode="PORT2"),
            ]
            commentary = {
                "PORT1": {"A": make_commentary_result("A", "Commentary A.")},
                "PORT2": {"B": make_commentary_result("B", "Commentary B.")},
            }
            
            result_path = create_output_workbook(selections, commentary, output_folder)
            
            # Verify sheet names
            wb = openpyxl.load_workbook(result_path)
            assert "PORT1" in wb.sheetnames
            assert "PORT2" in wb.sheetnames
            wb.close()

    def test_duplicate_ticker_across_portfolios_stays_isolated_by_sheet(self):
        """Duplicate tickers in different portfolios should not cross-populate output rows."""
        import openpyxl

        with tempfile.TemporaryDirectory() as tmpdir:
            output_folder = Path(tmpdir)

            selections = [
                make_selection_result([make_ranked_security("AAPL", 0.10)], portcode="XYZ"),
                make_selection_result([make_ranked_security("AAPL", 0.20)], portcode="ONE"),
            ]
            commentary = {
                "XYZ": {"AAPL": make_commentary_result("AAPL", "XYZ commentary.")},
                "ONE": {"AAPL": make_commentary_result("AAPL", "ONE commentary.")},
            }

            result_path = create_output_workbook(selections, commentary, output_folder)

            wb = openpyxl.load_workbook(result_path)
            ws_xyz = wb["XYZ"]
            ws_one = wb["ONE"]

            assert ws_xyz["A2"].value == "AAPL"
            assert ws_xyz["G2"].value == "XYZ commentary."
            assert ws_one["A2"].value == "AAPL"
            assert ws_one["G2"].value == "ONE commentary."
            assert ws_xyz["G2"].value != "ERROR: No commentary generated"
            assert ws_one["G2"].value != "ERROR: No commentary generated"
            wb.close()

    def test_creates_output_folder_if_needed(self):
        """Should create output folder if it doesn't exist."""
        with tempfile.TemporaryDirectory() as tmpdir:
            output_folder = Path(tmpdir) / "nested" / "folder"
            
            ranked = [make_ranked_security("AAPL", 0.15)]
            selection = make_selection_result(ranked)
            commentary = {"TEST": {"AAPL": make_commentary_result("AAPL", "Commentary.")}}
            
            result_path = create_output_workbook([selection], commentary, output_folder)
            
            assert output_folder.exists()
            assert result_path.exists()

    def test_writes_overview_table_above_security_table(self):
        """Should write overview table in rows 1-2 and shift security table down."""
        import openpyxl

        with tempfile.TemporaryDirectory() as tmpdir:
            output_folder = Path(tmpdir)

            ranked = [make_ranked_security("AAPL", 0.15)]
            selection = make_selection_result(ranked, portcode="PORT1")
            commentary = {"PORT1": {"AAPL": make_commentary_result("AAPL", "Security commentary.")}}
            overview = {
                "PORT1": make_attribution_overview_result(
                    portcode="PORT1",
                    output="Portfolio attribution overview text.",
                    citations=[Citation(url="https://reuters.com/overview")],
                )
            }

            result_path = create_output_workbook(
                [selection],
                commentary,
                output_folder,
                attribution_overview_results=overview,
            )

            wb = openpyxl.load_workbook(result_path)
            ws = wb["PORT1"]

            assert ws["A1"].value == "Category"
            assert ws["B1"].value == "Output"
            assert ws["C1"].value == "Sources"
            assert ws["A2"].value == "overview"
            assert "Portfolio attribution overview text." in ws["B2"].value
            assert "[1] https://reuters.com/overview" in ws["C2"].value

            # Security table should start at row 4 when overview is present.
            assert ws["A4"].value == "Ticker"
            assert ws["A5"].value == "AAPL"
            wb.close()

    def test_no_overview_keeps_legacy_security_header_row(self):
        """Should preserve legacy layout when overview results are not provided."""
        import openpyxl

        with tempfile.TemporaryDirectory() as tmpdir:
            output_folder = Path(tmpdir)

            ranked = [make_ranked_security("AAPL", 0.15)]
            selection = make_selection_result(ranked, portcode="PORT1")
            commentary = {"PORT1": {"AAPL": make_commentary_result("AAPL", "Security commentary.")}}

            result_path = create_output_workbook([selection], commentary, output_folder)

            wb = openpyxl.load_workbook(result_path)
            ws = wb["PORT1"]
            assert ws["A1"].value == "Ticker"
            assert ws["A2"].value == "AAPL"
            wb.close()

    def test_overview_warning_row_writes_error_text_and_empty_sources(self):
        """Failed overview results should render warning text with empty sources."""
        import openpyxl

        with tempfile.TemporaryDirectory() as tmpdir:
            output_folder = Path(tmpdir)

            ranked = [make_ranked_security("AAPL", 0.15)]
            selection = make_selection_result(ranked, portcode="PORT1")
            commentary = {"PORT1": {"AAPL": make_commentary_result("AAPL", "Security commentary.")}}
            overview = {
                "PORT1": make_attribution_overview_result(
                    portcode="PORT1",
                    success=False,
                    error_message="WARNING: No attribution data available.",
                )
            }

            result_path = create_output_workbook(
                [selection],
                commentary,
                output_folder,
                attribution_overview_results=overview,
            )

            wb = openpyxl.load_workbook(result_path)
            ws = wb["PORT1"]
            assert ws["B2"].value == "WARNING: No attribution data available."
            assert ws["C2"].value in ("", None)
            wb.close()


# --- create_log_file Tests ---

class TestCreateLogFile:
    """Tests for the create_log_file function."""

    def test_creates_log_file(self):
        """Should create a log file."""
        with tempfile.TemporaryDirectory() as tmpdir:
            output_folder = Path(tmpdir)
            input_files = [Path("file1.xlsx"), Path("file2.xlsx")]
            output_file = Path("output.xlsx")
            errors = {}
            start_time = datetime(2026, 1, 28, 10, 0, 0)
            end_time = datetime(2026, 1, 28, 10, 5, 30)
            
            log_path = create_log_file(
                output_folder, input_files, output_file, errors, start_time, end_time
            )
            
            assert log_path.exists()
            assert log_path.parent.name == "log"
            assert "run_log_" in log_path.name

    def test_log_contains_run_info(self):
        """Should include run information in log."""
        with tempfile.TemporaryDirectory() as tmpdir:
            output_folder = Path(tmpdir)
            input_files = [Path("portfolio1.xlsx")]
            output_file = Path("output.xlsx")
            start_time = datetime(2026, 1, 28, 10, 0, 0)
            end_time = datetime(2026, 1, 28, 10, 5, 30)
            
            log_path = create_log_file(
                output_folder, input_files, output_file, {}, start_time, end_time
            )
            
            content = log_path.read_text()
            assert "portfolio1.xlsx" in content
            assert "output.xlsx" in content
            assert "330.0 seconds" in content  # 5 min 30 sec
            assert "No errors encountered" in content

    def test_log_contains_errors(self):
        """Should include errors in log."""
        with tempfile.TemporaryDirectory() as tmpdir:
            output_folder = Path(tmpdir)
            errors = {
                "PORT1|AAPL": ["API timeout", "Retry failed"],
                "PORT1|MSFT": ["Invalid response"],
            }
            start_time = datetime(2026, 1, 28, 10, 0, 0)
            end_time = datetime(2026, 1, 28, 10, 1, 0)
            
            log_path = create_log_file(
                output_folder, [], Path("out.xlsx"), errors, start_time, end_time
            )
            
            content = log_path.read_text()
            assert "PORT1|AAPL" in content
            assert "API timeout" in content
            assert "Retry failed" in content
            assert "Invalid response" in content

    def test_creates_log_subfolder(self):
        """Should create log subfolder under output folder."""
        with tempfile.TemporaryDirectory() as tmpdir:
            output_folder = Path(tmpdir)
            start_time = datetime(2026, 1, 28, 10, 0, 0)
            
            log_path = create_log_file(
                output_folder, [], Path("out.xlsx"), {}, start_time, start_time
            )
            
            assert (output_folder / "log").is_dir()
            assert log_path.parent == output_folder / "log"
